"""SQL Editor panel — syntax highlighting, autocomplete, EXPLAIN, history."""

from __future__ import annotations
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QTableWidget, QTableWidgetItem, QPushButton,
    QLabel, QComboBox, QFileDialog, QMessageBox, QHeaderView,
    QAbstractItemView, QCompleter, QPlainTextEdit, QFrame,
    QDialog, QDialogButtonBox, QListWidget, QListWidgetItem,
    QLineEdit, QTextEdit as _QTextEdit,
)
from PyQt6.QtGui import (
    QFont, QSyntaxHighlighter, QTextCharFormat, QColor,
    QKeySequence, QShortcut, QTextCursor,
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QRegularExpression,
    QRect,
)

from ...db.connection import DBManager
from ...utils import i18n, config


# ── Syntax highlighter ─────────────────────────────────────────────────────

class SQLHighlighter(QSyntaxHighlighter):
    KEYWORDS = (
        "SELECT FROM WHERE JOIN LEFT RIGHT INNER OUTER FULL ON AS AND OR NOT "
        "IN IS NULL LIKE ILIKE BETWEEN ORDER BY GROUP HAVING LIMIT OFFSET "
        "INSERT INTO VALUES UPDATE SET DELETE CREATE TABLE INDEX VIEW DROP "
        "ALTER ADD COLUMN SCHEMA DATABASE WITH UNION ALL DISTINCT CASE WHEN "
        "THEN ELSE END EXISTS RETURNING TRUNCATE CASCADE VACUUM ANALYZE "
        "EXPLAIN BEGIN COMMIT ROLLBACK SAVEPOINT RELEASE TRUE FALSE LATERAL "
        "CROSS NATURAL USING OVER PARTITION WINDOW FILTER RECURSIVE"
    ).split()

    POSTGIS = (
        "ST_GeomFromText ST_AsText ST_AsGeoJSON ST_Transform ST_SRID "
        "ST_Area ST_Length ST_Perimeter ST_Distance ST_Buffer "
        "ST_Intersection ST_Union ST_Difference ST_Contains ST_Within "
        "ST_Overlaps ST_Touches ST_Intersects ST_DWithin ST_Envelope "
        "ST_ConvexHull ST_Centroid ST_Simplify ST_IsValid ST_IsSimple "
        "ST_IsEmpty ST_MakeValid ST_SetSRID ST_X ST_Y ST_Z "
        "ST_NPoints ST_GeometryN ST_ExteriorRing geometry_columns "
        "ST_LineInterpolatePoint ST_MakeLine ST_MakePoint "
        "postgis_version postgis_lib_version pgr_dijkstra "
        "pgr_drivingDistance pgr_version ST_SnapToGrid ST_Collect "
        "ST_GeoHash ST_AsEWKT ST_AsEWKB ST_GeomFromEWKT ST_Force2D "
        "ST_FlipCoordinates ST_GeneratePoints ST_VoronoiPolygons "
        "ST_ClosestPoint ST_ShortestLine ST_LongestLine ST_Expand"
    ).split()

    def __init__(self, doc):
        super().__init__(doc)
        self.rules: list[tuple] = []

        kw_fmt = QTextCharFormat()
        kw_fmt.setForeground(QColor("#569CD6"))
        kw_fmt.setFontWeight(700)
        for kw in self.KEYWORDS:
            self.rules.append((
                QRegularExpression(r'\b' + kw + r'\b',
                    QRegularExpression.PatternOption.CaseInsensitiveOption),
                kw_fmt))

        pg_fmt = QTextCharFormat()
        pg_fmt.setForeground(QColor("#4EC9B0"))
        for fn in self.POSTGIS:
            self.rules.append((
                QRegularExpression(r'\b' + fn + r'\b',
                    QRegularExpression.PatternOption.CaseInsensitiveOption),
                pg_fmt))

        str_fmt = QTextCharFormat()
        str_fmt.setForeground(QColor("#CE9178"))
        self.rules.append((QRegularExpression(r"'[^']*'"), str_fmt))

        num_fmt = QTextCharFormat()
        num_fmt.setForeground(QColor("#B5CEA8"))
        self.rules.append((QRegularExpression(r'\b\d+\.?\d*\b'), num_fmt))

        cmt_fmt = QTextCharFormat()
        cmt_fmt.setForeground(QColor("#6A9955"))
        cmt_fmt.setFontItalic(True)
        self.rules.append((QRegularExpression(r'--[^\n]*'), cmt_fmt))

        id_fmt = QTextCharFormat()
        id_fmt.setForeground(QColor("#DCDCAA"))
        self.rules.append((QRegularExpression(r'"[^"]+"'), id_fmt))

    def highlightBlock(self, text: str):
        for pattern, fmt in self.rules:
            it = pattern.globalMatch(text)
            while it.hasNext():
                m = it.next()
                self.setFormat(m.capturedStart(), m.capturedLength(), fmt)


# ── SQL Editor widget with autocomplete ────────────────────────────────────

class SQLEditor(QPlainTextEdit):
    """QPlainTextEdit with Ctrl+Space autocomplete popup."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFont(QFont("Courier New", 12))
        self.setPlaceholderText("-- Enter SQL here  (F5 to run, Ctrl+Space to autocomplete)")
        self._completer: QCompleter | None = None
        self._setup_completer([])

    def _setup_completer(self, words: list[str]):
        if self._completer:
            self._completer.setParent(None)
        all_words = sorted(set(
            SQLHighlighter.KEYWORDS + SQLHighlighter.POSTGIS + words
        ), key=str.lower)
        self._completer = QCompleter(all_words, self)
        self._completer.setWidget(self)
        self._completer.setCompletionMode(
            QCompleter.CompletionMode.PopupCompletion)
        self._completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._completer.activated.connect(self._insert_completion)

    def set_completions(self, words: list[str]):
        self._setup_completer(words)

    def _insert_completion(self, text: str):
        cur = self.textCursor()
        extra = len(text) - len(self._completer.completionPrefix())
        cur.movePosition(QTextCursor.MoveOperation.Left)
        cur.movePosition(QTextCursor.MoveOperation.EndOfWord)
        cur.insertText(text[-extra:])
        self.setTextCursor(cur)

    def _word_under_cursor(self) -> str:
        cur = self.textCursor()
        cur.select(QTextCursor.SelectionType.WordUnderCursor)
        return cur.selectedText()

    def keyPressEvent(self, event):
        popup = self._completer.popup()
        if popup.isVisible():
            if event.key() in (
                Qt.Key.Key_Enter, Qt.Key.Key_Return,
                Qt.Key.Key_Escape, Qt.Key.Key_Tab,
                Qt.Key.Key_Backtab,
            ):
                event.ignore()
                return
        # Ctrl+Space → trigger manually
        if (event.key() == Qt.Key.Key_Space and
                event.modifiers() == Qt.KeyboardModifier.ControlModifier):
            self._trigger_complete()
            return
        super().keyPressEvent(event)
        # Auto-trigger after typing ≥2 chars
        if event.text() and event.text().isalnum():
            self._trigger_complete()
        else:
            self._completer.popup().hide()

    def _trigger_complete(self):
        prefix = self._word_under_cursor()
        if len(prefix) < 2:
            self._completer.popup().hide()
            return
        if prefix != self._completer.completionPrefix():
            self._completer.setCompletionPrefix(prefix)
            self._completer.popup().setCurrentIndex(
                self._completer.completionModel().index(0, 0))
        cur_rect: QRect = self.cursorRect()
        cur_rect.setWidth(
            self._completer.popup().sizeHintForColumn(0)
            + self._completer.popup().verticalScrollBar().sizeHint().width())
        self._completer.complete(cur_rect)


# ── Workers ────────────────────────────────────────────────────────────────

class QueryWorker(QThread):
    done  = pyqtSignal(object, object, float)
    error = pyqtSignal(str)

    def __init__(self, db: DBManager, sql: str):
        super().__init__()
        self.db = db
        self.sql = sql

    def run(self):
        try:
            cols, rows, ms = self.db.execute_sql(self.sql)
            self.done.emit(cols, rows, ms)
        except Exception as e:
            self.error.emit(str(e))


class CompletionLoader(QThread):
    """Load schema/table/column/function names for autocomplete."""
    done = pyqtSignal(list)

    def __init__(self, db: DBManager):
        super().__init__()
        self.db = db

    def run(self):
        words: list[str] = []
        try:
            import psycopg2
            conn = psycopg2.connect(**self.db.params)
            cur = conn.cursor()
            # schemas
            cur.execute("SELECT schema_name FROM information_schema.schemata")
            words += [r[0] for r in cur.fetchall()]
            # tables + columns
            cur.execute("""
                SELECT table_schema, table_name, column_name
                FROM information_schema.columns
                WHERE table_schema NOT IN ('pg_catalog','information_schema')
                LIMIT 5000
            """)
            for schema, table, col in cur.fetchall():
                words += [table, f'"{schema}"."{table}"', col]
            # pg functions (public + postgis)
            cur.execute("""
                SELECT routine_name
                FROM information_schema.routines
                WHERE routine_schema IN ('public','postgis')
                LIMIT 1000
            """)
            words += [r[0] for r in cur.fetchall()]
            cur.close()
            conn.close()
        except Exception:
            pass
        self.done.emit(list(set(words)))


# ── Main panel ─────────────────────────────────────────────────────────────

# ── Geometry column detection ──────────────────────────────────────────────

_GEO_NAMES = frozenset({
    "geom", "geometry", "the_geom", "shape", "wkb_geometry",
    "geomfromtext", "geo", "geog", "geography", "point",
    "line", "polygon", "multipolygon", "multilinestring", "multipoint",
})

def _is_geo_col(col_name: str, sample_value) -> bool:
    """Return True if this column looks like a PostGIS geometry."""
    if col_name.lower() in _GEO_NAMES:
        return True
    if col_name.lower().startswith("st_"):
        return True
    if isinstance(sample_value, str) and len(sample_value) > 20:
        stripped = sample_value.strip()
        # WKB hex — all hex chars, typically starts with 0 or 1
        if all(c in "0123456789abcdefABCDEF" for c in stripped[:20]):
            return True
        # WKT
        if any(stripped.upper().startswith(k) for k in (
                "POINT", "LINESTRING", "POLYGON", "MULTI",
                "GEOMETRYCOLLECTION", "SRID=")):
            return True
    return False


class GeoQueryWorker(QThread):
    """Re-runs SQL wrapping geometry cols with ST_AsGeoJSON."""
    done  = pyqtSignal(list, list)   # features: [(geojson_str, attrs)], col_names
    error = pyqtSignal(str)

    def __init__(self, db: DBManager, sql: str, geo_cols: list[str], attr_cols: list[str]):
        super().__init__()
        self.db       = db
        self.sql      = sql
        self.geo_cols = geo_cols
        self.attr_cols = attr_cols

    def run(self):
        try:
            import psycopg2
            conn = psycopg2.connect(**self.db.params,
                                    options="-c client_encoding=UTF8")
            cur  = conn.cursor()

            # Build SELECT: ST_AsGeoJSON for each geo col, rest as-is
            selects = []
            for col in self.geo_cols:
                selects.append(f'ST_AsGeoJSON("{col}", 6) AS "_geojson_{col}"')
            for col in self.attr_cols:
                selects.append(f'"{col}"')

            wrapped = (f"SELECT {', '.join(selects)} "
                       f"FROM ({self.sql}) AS _sql_result LIMIT 50000")
            cur.execute(wrapped)
            desc  = [d[0] for d in cur.description]
            rows  = cur.fetchall()
            cur.close()
            conn.close()

            features = []
            for row in rows:
                rd = dict(zip(desc, row))
                # one feature per geometry column (use first geo col)
                geojson_str = rd.get(f"_geojson_{self.geo_cols[0]}")
                if geojson_str is None:
                    continue
                attrs = {c: rd.get(c) for c in self.attr_cols}
                features.append((geojson_str, attrs))

            self.done.emit(features, self.attr_cols)
        except Exception as e:
            self.error.emit(str(e))


class SQLEditorPanel(QWidget):
    # Emitted when user clicks "Show on Map"
    # payload: list of (geojson_str, attrs_dict), list of attr col names
    show_on_map = pyqtSignal(list, list, str)

    def __init__(self, db: DBManager, parent=None):
        super().__init__(parent)
        self.db = db
        raw_hist = config.get("sql_history", [])
        # migrate old plain-string entries to {sql, ts} dicts
        self._history: list[dict] = [
            h if isinstance(h, dict) else {"sql": h, "ts": "", "pinned": False}
            for h in raw_hist
        ]
        self._worker: QueryWorker | None = None
        self._geo_worker: GeoQueryWorker | None = None
        self._comp_loader: CompletionLoader | None = None
        self._active_schema = ""
        self._active_table = ""
        self._last_sql = ""
        self._detected_geo_cols: list[str] = []
        self._detected_attr_cols: list[str] = []
        self._build_ui()
        for entry in self._history:
            sql = entry["sql"] if isinstance(entry, dict) else entry
            self._history_combo.addItem(sql[:60].replace("\n", " "))

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        # ── Toolbar ──────────────────────────────────────────────────────
        toolbar = QHBoxLayout()
        self._run_btn = QPushButton(i18n.t("sql_run"))
        self._run_btn.clicked.connect(self._run_query)
        toolbar.addWidget(self._run_btn)

        self._stop_btn = QPushButton(i18n.t("sql_stop"))
        self._stop_btn.setEnabled(False)
        self._stop_btn.clicked.connect(self._stop_query)
        toolbar.addWidget(self._stop_btn)

        self._explain_btn = QPushButton("⚡ EXPLAIN")
        self._explain_btn.setToolTip("Run EXPLAIN ANALYZE and show visual plan")
        self._explain_btn.clicked.connect(self._run_explain)
        toolbar.addWidget(self._explain_btn)

        self._clear_btn = QPushButton(i18n.t("sql_clear"))
        self._clear_btn.clicked.connect(self._clear)
        toolbar.addWidget(self._clear_btn)

        toolbar.addStretch()

        self._save_btn = QPushButton(i18n.t("sql_save"))
        self._save_btn.clicked.connect(self._save_query)
        toolbar.addWidget(self._save_btn)

        self._load_btn = QPushButton(i18n.t("sql_load"))
        self._load_btn.clicked.connect(self._load_query)
        toolbar.addWidget(self._load_btn)

        self._history_combo = QComboBox()
        self._history_combo.setMinimumWidth(200)
        self._history_combo.setPlaceholderText("History...")
        self._history_combo.currentIndexChanged.connect(self._load_history)
        toolbar.addWidget(self._history_combo)
        layout.addLayout(toolbar)

        # ── Splitter: editor | results ────────────────────────────────────
        splitter = QSplitter(Qt.Orientation.Vertical)
        layout.addWidget(splitter)

        # Editor
        editor_widget = QWidget()
        ed_layout = QVBoxLayout(editor_widget)
        ed_layout.setContentsMargins(0, 0, 0, 0)

        tpl_row = QHBoxLayout()
        for name, sql in [
            ("SELECT *",  'SELECT * FROM "{schema}"."{table}" LIMIT 100;'),
            ("COUNT",     'SELECT COUNT(*) FROM "{schema}"."{table}";'),
            ("Extent",    'SELECT ST_Extent(geom) FROM "{schema}"."{table}";'),
            ("Validate",  'SELECT ctid, ST_IsValidReason(geom) FROM "{schema}"."{table}" WHERE NOT ST_IsValid(geom);'),
            ("Columns",   "SELECT column_name, udt_name FROM information_schema.columns WHERE table_schema='{schema}' AND table_name='{table}';"),
        ]:
            btn = QPushButton(name)
            btn.setFixedHeight(24)
            btn.clicked.connect(lambda _, s=sql: self._insert_template(s))
            tpl_row.addWidget(btn)
        tpl_row.addStretch()
        ed_layout.addLayout(tpl_row)

        self._editor = SQLEditor()
        self._highlighter = SQLHighlighter(self._editor.document())
        ed_layout.addWidget(self._editor)
        splitter.addWidget(editor_widget)

        # Results tab
        result_widget = QWidget()
        res_layout = QVBoxLayout(result_widget)
        res_layout.setContentsMargins(0, 4, 0, 0)

        # result status + map button row
        status_row = QHBoxLayout()
        self._result_label = QLabel("")
        status_row.addWidget(self._result_label, 1)

        self._map_banner = QFrame()
        self._map_banner.setStyleSheet(
            "QFrame{background:#1a3a1a;border:1px solid #2e7d32;"
            "border-radius:4px;padding:2px;}")
        banner_lay = QHBoxLayout(self._map_banner)
        banner_lay.setContentsMargins(8, 2, 4, 2)
        self._map_banner_lbl = QLabel("🗺  Geometry column detected:")
        self._map_banner_lbl.setStyleSheet("color:#81c784;font-weight:bold;")
        self._btn_show_map = QPushButton("Show on Map")
        self._btn_show_map.setStyleSheet(
            "QPushButton{background:#2e7d32;color:#fff;border:none;"
            "border-radius:3px;padding:2px 10px;font-weight:bold;}"
            "QPushButton:hover{background:#388e3c;}"
            "QPushButton:disabled{background:#444;color:#777;}")
        self._btn_show_map.clicked.connect(self._request_show_on_map)
        banner_lay.addWidget(self._map_banner_lbl)
        banner_lay.addWidget(self._btn_show_map)
        self._map_banner.hide()
        status_row.addWidget(self._map_banner)

        res_layout.addLayout(status_row)

        self._result_table = QTableWidget()
        self._result_table.setAlternatingRowColors(True)
        self._result_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self._result_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self._result_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        # Column header right-click → stats
        self._result_table.horizontalHeader().setContextMenuPolicy(
            Qt.ContextMenuPolicy.CustomContextMenu)
        self._result_table.horizontalHeader().customContextMenuRequested.connect(
            self._show_result_col_stats)
        res_layout.addWidget(self._result_table)

        exp_row = QHBoxLayout()
        self._export_btn = QPushButton("💾 Export CSV")
        self._export_btn.setToolTip("Export results to CSV file")
        self._export_btn.clicked.connect(self._export_csv)
        exp_row.addWidget(self._export_btn)

        self._export_xlsx_btn = QPushButton("📊 Export Excel")
        self._export_xlsx_btn.setToolTip("Export results to Excel .xlsx file")
        self._export_xlsx_btn.clicked.connect(self._export_xlsx)
        exp_row.addWidget(self._export_xlsx_btn)

        self._fmt_btn = QPushButton("✨ Format SQL")
        self._fmt_btn.setToolTip("Auto-format / pretty-print the SQL query")
        self._fmt_btn.clicked.connect(self._format_sql)
        exp_row.addWidget(self._fmt_btn)

        self._history_btn = QPushButton("🕐 History")
        self._history_btn.setToolTip("Browse query history with timestamps")
        self._history_btn.clicked.connect(self._show_history_panel)
        exp_row.addWidget(self._history_btn)

        exp_row.addStretch()
        res_layout.addLayout(exp_row)
        splitter.addWidget(result_widget)
        splitter.setSizes([300, 300])

        QShortcut(QKeySequence("F5"), self, self._run_query)
        QShortcut(QKeySequence("F6"), self, self._run_explain)

    # ── Autocomplete ──────────────────────────────────────────────────────

    def refresh_completions(self):
        if not self.db.is_connected():
            return
        if self._comp_loader and self._comp_loader.isRunning():
            return
        self._comp_loader = CompletionLoader(self.db)
        self._comp_loader.done.connect(self._editor.set_completions)
        self._comp_loader.start()

    # ── Layer context ─────────────────────────────────────────────────────

    def set_active_layer(self, schema: str, table: str):
        self._active_schema = schema
        self._active_table = table

    def _insert_template(self, sql: str):
        sql = sql.replace("{schema}", self._active_schema or "public")
        sql = sql.replace("{table}", self._active_table or "layer_name")
        self._editor.setPlainText(sql)

    # ── Query execution ───────────────────────────────────────────────────

    def _run_query(self):
        if not self.db.is_connected():
            QMessageBox.warning(self, "Error", i18n.t("err_not_connected"))
            return
        sql = self._editor.toPlainText().strip()
        if not sql:
            return
        self._run_btn.setEnabled(False)
        self._stop_btn.setEnabled(True)
        self._result_label.setText("Running...")
        import datetime as _dt
        first_sql = self._history[0]["sql"] if self._history else None
        if first_sql != sql:
            entry = {"sql": sql,
                     "ts": _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     "pinned": False}
            self._history.insert(0, entry)
            self._history = self._history[:100]
            self._history_combo.insertItem(0, sql[:60].replace("\n", " "))
            config.set("sql_history", self._history)
        self._worker = QueryWorker(self.db, sql)
        self._worker.done.connect(self._on_result)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_result(self, cols, rows, ms: float):
        self._run_btn.setEnabled(True)
        self._stop_btn.setEnabled(False)
        self._map_banner.hide()
        self._detected_geo_cols = []
        self._detected_attr_cols = []

        if cols and rows is not None:
            n = len(rows)
            self._result_label.setText(
                i18n.t("sql_rows_returned", n=n, ms=f"{ms:.1f}"))
            self._result_table.setRowCount(n)
            self._result_table.setColumnCount(len(cols))
            self._result_table.setHorizontalHeaderLabels(cols)

            # detect geometry columns from first non-null row
            first_vals = {}
            for row in rows:
                for c, col in enumerate(cols):
                    if row[c] is not None and col not in first_vals:
                        first_vals[col] = row[c]
                if len(first_vals) == len(cols):
                    break

            geo_cols  = []
            attr_cols = []
            for col in cols:
                sample = first_vals.get(col)
                if _is_geo_col(col, sample):
                    geo_cols.append(col)
                else:
                    attr_cols.append(col)

            for r, row in enumerate(rows):
                for c, val in enumerate(row):
                    display = str(val) if val is not None else ""
                    # truncate WKB blobs for display
                    if len(display) > 60 and cols[c] in geo_cols:
                        display = display[:30] + "…"
                    self._result_table.setItem(r, c, QTableWidgetItem(display))

            if geo_cols:
                self._detected_geo_cols  = geo_cols
                self._detected_attr_cols = attr_cols
                names = ", ".join(geo_cols)
                self._map_banner_lbl.setText(f"🗺  Geometry detected: {names}")
                self._map_banner.show()
        else:
            self._result_label.setText(
                i18n.t("sql_no_result") + f"  ({ms:.1f} ms)")
            self._result_table.setRowCount(0)
            self._result_table.setColumnCount(0)
        if self.parent() and hasattr(self.parent(), "log"):
            self.parent().log(f"SQL done in {ms:.1f} ms", "success")

    def _on_error(self, error: str):
        self._run_btn.setEnabled(True)
        self._stop_btn.setEnabled(False)
        self._result_label.setText(f"Error: {error}")
        if self.parent() and hasattr(self.parent(), "log"):
            self.parent().log(f"SQL Error: {error}", "error")

    def _request_show_on_map(self):
        if not self._detected_geo_cols:
            return
        sql = self._editor.toPlainText().strip()
        if not sql:
            return
        self._btn_show_map.setEnabled(False)
        self._btn_show_map.setText("Loading…")
        label = sql[:60].replace("\n", " ")
        w = GeoQueryWorker(self.db, sql,
                           self._detected_geo_cols,
                           self._detected_attr_cols)
        w.done.connect(lambda feats, cols:
                       self._on_geo_ready(feats, cols, label))
        w.error.connect(self._on_geo_error)
        self._geo_worker = w
        w.start()

    def _on_geo_ready(self, features: list, attr_cols: list, label: str):
        self._btn_show_map.setEnabled(True)
        self._btn_show_map.setText("Show on Map")
        self.show_on_map.emit(features, attr_cols, label)

    def _on_geo_error(self, error: str):
        self._btn_show_map.setEnabled(True)
        self._btn_show_map.setText("Show on Map")
        self._result_label.setText(f"Map error: {error}")

    def _stop_query(self):
        if self._worker and self._worker.isRunning():
            self._worker.terminate()
        self._run_btn.setEnabled(True)
        self._stop_btn.setEnabled(False)

    def _clear(self):
        self._editor.clear()
        self._result_table.setRowCount(0)
        self._result_table.setColumnCount(0)
        self._result_label.setText("")

    # ── EXPLAIN ───────────────────────────────────────────────────────────

    def _run_explain(self):
        if not self.db.is_connected():
            QMessageBox.warning(self, "Error", i18n.t("err_not_connected"))
            return
        sql = self._editor.toPlainText().strip()
        if not sql:
            return
        from ..dialogs.explain_dialog import ExplainDialog
        ExplainDialog(self.db, sql, self).exec()

    # ── Save / load ───────────────────────────────────────────────────────

    def _save_query(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "Save Query", "Query name:")
        if ok and name:
            config.save_query(name, self._editor.toPlainText())

    def _load_query(self):
        from PyQt6.QtWidgets import QInputDialog
        queries = config.get_saved_queries()
        if not queries:
            QMessageBox.information(self, "Info", "No saved queries.")
            return
        names = [q["name"] for q in queries]
        name, ok = QInputDialog.getItem(
            self, "Load Query", "Select:", names, 0, False)
        if ok:
            q = next((q for q in queries if q["name"] == name), None)
            if q:
                self._editor.setPlainText(q["sql"])

    def _load_history(self, idx: int):
        if 0 <= idx < len(self._history):
            entry = self._history[idx]
            sql = entry["sql"] if isinstance(entry, dict) else entry
            self._editor.setPlainText(sql)

    # ── Export ────────────────────────────────────────────────────────────

    def _export_csv(self):
        import csv as csv_mod
        path, _ = QFileDialog.getSaveFileName(
            self, "Export", "result.csv", "CSV (*.csv)")
        if not path:
            return
        rows = self._result_table.rowCount()
        cols = self._result_table.columnCount()
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv_mod.writer(f)
            w.writerow([self._result_table.horizontalHeaderItem(c).text()
                        for c in range(cols)])
            for r in range(rows):
                w.writerow([
                    self._result_table.item(r, c).text()
                    if self._result_table.item(r, c) else ""
                    for c in range(cols)
                ])
        QMessageBox.information(self, "Done", f"Exported to {path}")

    def _export_xlsx(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(
                self, "Missing dependency",
                "openpyxl is required for Excel export.\n\n"
                "Install with:  pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", "result.xlsx",
            "Excel Workbook (*.xlsx)")
        if not path:
            return
        rows = self._result_table.rowCount()
        cols = self._result_table.columnCount()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Query Results"

        hdr_fill = PatternFill("solid", fgColor="2E4057")
        hdr_font = Font(bold=True, color="FFFFFF")
        for c in range(cols):
            h = self._result_table.horizontalHeaderItem(c)
            cell = ws.cell(row=1, column=c + 1,
                           value=h.text() if h else f"col{c}")
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for r in range(rows):
            for c in range(cols):
                item = self._result_table.item(r, c)
                val = item.text() if item else ""
                try:
                    val = float(val) if "." in val else int(val)
                except (ValueError, TypeError):
                    pass
                ws.cell(row=r + 2, column=c + 1, value=val)

        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 2, 40)

        wb.save(path)
        QMessageBox.information(self, "Done", f"Exported to {path}")

    # ── Format SQL ────────────────────────────────────────────────────────

    def _format_sql(self):
        sql = self._editor.toPlainText().strip()
        if not sql:
            return
        try:
            import sqlparse
            formatted = sqlparse.format(
                sql,
                reindent=True,
                keyword_case="upper",
                identifier_case="lower",
                strip_comments=False,
                indent_width=4,
            )
            self._editor.setPlainText(formatted)
            return
        except ImportError:
            pass
        # Fallback: simple keyword uppercasing without sqlparse
        import re
        keywords = (
            "select from where join left right inner outer on as and or not "
            "in is null like between order by group having limit offset "
            "insert into values update set delete create table index view drop "
            "with union all distinct case when then else end exists returning "
            "truncate cascade vacuum analyze explain begin commit rollback"
        ).split()
        result = sql
        for kw in keywords:
            result = re.sub(
                r'\b' + re.escape(kw) + r'\b', kw.upper(), result,
                flags=re.IGNORECASE)
        self._editor.setPlainText(result)

    # ── Result column stats ───────────────────────────────────────────────

    def _show_result_col_stats(self, pos):
        col = self._result_table.horizontalHeader().logicalIndexAt(pos)
        if col < 0:
            return
        h = self._result_table.horizontalHeaderItem(col)
        col_name = h.text() if h else f"col {col}"
        rows = self._result_table.rowCount()
        numeric, text_vals, null_n = [], [], 0
        for r in range(rows):
            item = self._result_table.item(r, col)
            val = item.text() if item else ""
            if not val or val in ("None", "NULL"):
                null_n += 1
            else:
                try:
                    numeric.append(float(val))
                except ValueError:
                    text_vals.append(val)
        lines = [
            f"<b>Column:</b> {col_name}",
            f"<b>Total rows:</b> {rows}",
            f"<b>Null / empty:</b> {null_n}",
            f"<b>Non-null:</b> {len(numeric) + len(text_vals)}",
        ]
        if numeric:
            avg = sum(numeric) / len(numeric)
            lines += [
                f"<b>Min:</b> {min(numeric):.6g}",
                f"<b>Max:</b> {max(numeric):.6g}",
                f"<b>Sum:</b> {sum(numeric):.6g}",
                f"<b>Average:</b> {avg:.6g}",
            ]
        if text_vals:
            lines.append(f"<b>Unique text values:</b> {len(set(text_vals))}")
        QMessageBox.information(
            self, f"Stats — {col_name}", "<br>".join(lines))

    # ── History panel ─────────────────────────────────────────────────────

    def _show_history_panel(self):
        dlg = _HistoryDialog(self._history, parent=self)
        if dlg.exec() and dlg.selected_sql():
            self._editor.setPlainText(dlg.selected_sql())
        self._history = dlg.history()
        config.set("sql_history", self._history)


# ── History Dialog ────────────────────────────────────────────────────────

class _HistoryDialog(QDialog):
    """Full history browser: search, pin, delete, preview."""

    def __init__(self, history: list[dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Query History")
        self.resize(820, 520)
        self._history = [
            h if isinstance(h, dict)
            else {"sql": h, "ts": "", "pinned": False}
            for h in history
        ]
        self._selected_sql: str = ""
        self._build_ui()
        self._populate()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Search bar
        search_row = QHBoxLayout()
        self._search = QLineEdit()
        self._search.setPlaceholderText("🔍 Search history…")
        self._search.textChanged.connect(self._filter)
        search_row.addWidget(self._search, 1)
        pin_btn = QPushButton("📌 Pin selected")
        pin_btn.clicked.connect(self._pin_selected)
        search_row.addWidget(pin_btn)
        del_btn = QPushButton("🗑 Delete selected")
        del_btn.clicked.connect(self._delete_selected)
        search_row.addWidget(del_btn)
        clear_btn = QPushButton("🗑 Clear all")
        clear_btn.clicked.connect(self._clear_all)
        search_row.addWidget(clear_btn)
        lay.addLayout(search_row)

        splitter = QSplitter(Qt.Orientation.Horizontal)

        # List
        self._list = QListWidget()
        self._list.currentRowChanged.connect(self._on_select)
        self._list.itemDoubleClicked.connect(lambda _: self.accept())
        splitter.addWidget(self._list)

        # Preview
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        self._ts_lbl = QLabel()
        self._ts_lbl.setStyleSheet("color:#888;font-size:11px;padding:2px 4px;")
        rl.addWidget(self._ts_lbl)
        self._preview = _QTextEdit()
        self._preview.setReadOnly(True)
        self._preview.setFont(QFont("Courier New", 10))
        rl.addWidget(self._preview)
        splitter.addWidget(right)
        splitter.setSizes([300, 520])

        lay.addWidget(splitter, 1)

        bb = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._use_selected)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)

    def _populate(self, query: str = ""):
        self._list.clear()
        self._visible_indices: list[int] = []
        for i, entry in enumerate(self._history):
            sql = entry.get("sql", "")
            if query and query.lower() not in sql.lower():
                continue
            pinned = entry.get("pinned", False)
            ts = entry.get("ts", "")
            preview = sql[:80].replace("\n", " ")
            label = ("📌 " if pinned else "") + preview
            item = QListWidgetItem(label)
            if pinned:
                item.setForeground(QColor("#f39c12"))
            self._list.addItem(item)
            self._visible_indices.append(i)

    def _filter(self, text: str):
        self._populate(text)

    def _on_select(self, row: int):
        if row < 0 or row >= len(self._visible_indices):
            return
        idx = self._visible_indices[row]
        entry = self._history[idx]
        self._preview.setPlainText(entry.get("sql", ""))
        ts = entry.get("ts", "")
        self._ts_lbl.setText(f"Executed: {ts}" if ts else "")

    def _pin_selected(self):
        row = self._list.currentRow()
        if row < 0 or row >= len(self._visible_indices):
            return
        idx = self._visible_indices[row]
        self._history[idx]["pinned"] = not self._history[idx].get("pinned", False)
        self._populate(self._search.text())

    def _delete_selected(self):
        row = self._list.currentRow()
        if row < 0 or row >= len(self._visible_indices):
            return
        idx = self._visible_indices[row]
        self._history.pop(idx)
        self._populate(self._search.text())

    def _clear_all(self):
        from PyQt6.QtWidgets import QMessageBox
        if QMessageBox.question(
                self, "Clear history", "Delete all query history?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            self._history = [h for h in self._history
                             if h.get("pinned", False)]
            self._populate(self._search.text())

    def _use_selected(self):
        row = self._list.currentRow()
        if row >= 0 and row < len(self._visible_indices):
            idx = self._visible_indices[row]
            self._selected_sql = self._history[idx].get("sql", "")
        self.accept()

    def selected_sql(self) -> str:
        return self._selected_sql

    def history(self) -> list[dict]:
        return self._history
